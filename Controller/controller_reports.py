
from datetime import datetime
from Models.base import get_connection
from Models.model_db import Database
import csv
import os

# Singleton database instance
_db = Database()


def get_full_election_report_data(election_id: int) -> dict:
    """
    Gather ALL raw data for a comprehensive election report.
    Returns dict with full voting records, candidates, voters info.
    """
    result = {
        "success": False,
        "error": None,
        "election": None,
        "positions": [],
        "candidates": [],
        "voting_records": [],
        "voters": [],
        "stats": {},
        "integrity": {},
        "generated_at": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    conn = get_connection()
    if not conn:
        result["error"] = "Database connection failed"
        return result

    try:
        cursor = conn.cursor(dictionary=True)

        # Get election info
        cursor.execute(
            """
            SELECT election_id, title, description, status, start_date, end_date,
                   allowed_grade, allowed_section
            FROM elections
            WHERE election_id = %s
            """,
            (election_id,),
        )
        election = cursor.fetchone()
        if not election:
            result["error"] = "Election not found"
            cursor.close()
            conn.close()
            return result

        result["election"] = election

        # Get positions for this election (ordered)
        cursor.execute(
            """
            SELECT position_id, election_id, title, display_order, created_at
            FROM positions
            WHERE election_id = %s
            ORDER BY display_order ASC, position_id ASC
            """,
            (election_id,),
        )
        result["positions"] = cursor.fetchall() or []

        # Get ALL candidates with full details
        cursor.execute(
            """
            SELECT c.candidate_id, c.election_id, c.position_id,
                   COALESCE(p.title, c.position, 'Unassigned') AS position_title,
                   c.full_name, c.slogan, c.bio, c.email, c.phone,
                   c.platform, c.photo_path, c.vote_count,
                   COALESCE(v.vote_total, 0) AS actual_votes
            FROM candidates c
            LEFT JOIN positions p ON p.position_id = c.position_id
            LEFT JOIN (
                SELECT candidate_id, COUNT(*) AS vote_total
                FROM voting_records
                WHERE candidate_id IS NOT NULL
                GROUP BY candidate_id
            ) v ON v.candidate_id = c.candidate_id
            WHERE c.election_id = %s
            ORDER BY position_title ASC, actual_votes DESC, c.full_name ASC
            """,
            (election_id,),
        )
        result["candidates"] = cursor.fetchall()

        # Get ALL voting records with voter and candidate details
        cursor.execute(
            """
            SELECT 
                vr.record_id,
                vr.user_id,
                u.username AS voter_username,
                u.full_name AS voter_name,
                u.student_id AS voter_student_id,
                u.email AS voter_email,
                u.grade_level AS voter_grade,
                u.section AS voter_section,
                vr.election_id,
                e.title AS election_title,
                vr.position_id,
                COALESCE(p.title, 'Unassigned') AS position_title,
                vr.candidate_id,
                c.full_name AS candidate_name,
                vr.status AS vote_status,
                vr.voted_at
            FROM voting_records vr
            LEFT JOIN users u ON u.user_id = vr.user_id
            LEFT JOIN elections e ON e.election_id = vr.election_id
            LEFT JOIN positions p ON p.position_id = vr.position_id
            LEFT JOIN candidates c ON c.candidate_id = vr.candidate_id
            WHERE vr.election_id = %s
            ORDER BY vr.voted_at DESC
            """,
            (election_id,),
        )
        result["voting_records"] = cursor.fetchall()

        # Get ALL voters who participated
        cursor.execute(
            """
            SELECT DISTINCT
                u.user_id,
                u.username,
                u.full_name,
                u.student_id,
                u.email,
                u.grade_level,
                u.section,
                u.role,
                u.created_at AS user_created_at,
                vr.voted_at
            FROM users u
            INNER JOIN voting_records vr ON vr.user_id = u.user_id
            WHERE vr.election_id = %s
            ORDER BY vr.voted_at DESC
            """,
            (election_id,),
        )
        result["voters"] = cursor.fetchall()

        # ------------------------------------------------------------------
        # STATS (turnout, cast/spoiled, eligibility)
        # ------------------------------------------------------------------
        allowed_grade = election.get("allowed_grade")
        allowed_section = (election.get("allowed_section") or "").strip()

        eligible_query = "SELECT COUNT(*) AS cnt FROM users WHERE role='student'"
        eligible_params = []
        if allowed_grade is not None:
            eligible_query += " AND grade_level = %s"
            eligible_params.append(allowed_grade)
        if allowed_section and allowed_section.upper() != "ALL":
            eligible_query += " AND UPPER(COALESCE(section,'')) = UPPER(%s)"
            eligible_params.append(allowed_section)

        cursor.execute(eligible_query, tuple(eligible_params))
        eligible_voters = int((cursor.fetchone() or {}).get("cnt") or 0)

        cursor.execute(
            """
            SELECT
                COUNT(*) AS total_records,
                SUM(CASE WHEN status='cast' THEN 1 ELSE 0 END) AS cast_records,
                SUM(CASE WHEN status='spoiled' THEN 1 ELSE 0 END) AS spoiled_records,
                COUNT(DISTINCT user_id) AS distinct_voters
            FROM voting_records
            WHERE election_id = %s
            """,
            (election_id,),
        )
        row = cursor.fetchone() or {}
        total_records = int(row.get("total_records") or 0)
        cast_records = int(row.get("cast_records") or 0)
        spoiled_records = int(row.get("spoiled_records") or 0)
        distinct_voters = int(row.get("distinct_voters") or 0)

        turnout_pct = (distinct_voters / eligible_voters * 100.0) if eligible_voters > 0 else 0.0
        result["stats"] = {
            "eligible_voters": eligible_voters,
            "participants": distinct_voters,
            "turnout_pct": turnout_pct,
            "total_records": total_records,
            "cast_records": cast_records,
            "spoiled_records": spoiled_records,
        }

        # Per-position summary
        cursor.execute(
            """
            SELECT
                p.position_id,
                p.title AS position_title,
                COUNT(vr.record_id) AS total_ballots,
                SUM(CASE WHEN vr.status='cast' THEN 1 ELSE 0 END) AS cast_ballots,
                SUM(CASE WHEN vr.status='spoiled' THEN 1 ELSE 0 END) AS spoiled_ballots,
                COUNT(DISTINCT vr.user_id) AS distinct_voters
            FROM positions p
            LEFT JOIN voting_records vr
                ON vr.position_id = p.position_id AND vr.election_id = p.election_id
            WHERE p.election_id = %s
            GROUP BY p.position_id, p.title
            ORDER BY p.display_order ASC, p.position_id ASC
            """,
            (election_id,),
        )
        result["stats"]["positions"] = cursor.fetchall() or []

        # ------------------------------------------------------------------
        # INTEGRITY CHECKS (professional audit counters)
        # ------------------------------------------------------------------
        cursor.execute(
            """
            SELECT COUNT(*) AS cnt
            FROM voting_records
            WHERE election_id=%s AND status='cast' AND (candidate_id IS NULL)
            """,
            (election_id,),
        )
        cast_missing_candidate = int((cursor.fetchone() or {}).get("cnt") or 0)

        cursor.execute(
            """
            SELECT COUNT(*) AS cnt
            FROM voting_records vr
            LEFT JOIN users u ON u.user_id = vr.user_id
            WHERE vr.election_id=%s AND u.user_id IS NULL
            """,
            (election_id,),
        )
        orphan_user_votes = int((cursor.fetchone() or {}).get("cnt") or 0)

        cursor.execute(
            """
            SELECT COUNT(*) AS cnt
            FROM voting_records vr
            LEFT JOIN candidates c ON c.candidate_id = vr.candidate_id
            WHERE vr.election_id=%s AND vr.candidate_id IS NOT NULL AND c.candidate_id IS NULL
            """,
            (election_id,),
        )
        orphan_candidate_votes = int((cursor.fetchone() or {}).get("cnt") or 0)

        cursor.execute(
            """
            SELECT COUNT(*) AS cnt
            FROM voting_records vr
            LEFT JOIN positions p ON p.position_id = vr.position_id
            WHERE vr.election_id=%s AND vr.position_id IS NOT NULL AND p.position_id IS NULL
            """,
            (election_id,),
        )
        orphan_position_votes = int((cursor.fetchone() or {}).get("cnt") or 0)

        result["integrity"] = {
            "cast_missing_candidate": cast_missing_candidate,
            "orphan_user_votes": orphan_user_votes,
            "orphan_candidate_votes": orphan_candidate_votes,
            "orphan_position_votes": orphan_position_votes,
        }

        result["success"] = True
        cursor.close()
        conn.close()

    except Exception as e:
        result["error"] = str(e)
        try:
            conn.close()
        except Exception:
            pass

    return result


def generate_pdf_report(report_data: dict, file_path: str) -> tuple[bool, str]:
    """
    Generate a highly polished, professional PDF report.
    Uses STRING headers to guarantee visibility and Paragraphs for data.
    """
    if not report_data.get("success"):
        return False, report_data.get("error", "No data available")

    # Import ReportLab modules
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                        TableStyle, Image, PageBreak, KeepTogether)
        from reportlab.lib.units import mm
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    except ImportError:
        return False, "reportlab library not installed. Install with: pip install reportlab"

    # Check for matplotlib
    have_mpl = True
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        from io import BytesIO
    except ImportError:
        have_mpl = False

    # Extract Data
    election = report_data.get("election", {})
    positions = report_data.get("positions", [])
    candidates = report_data.get("candidates", [])
    records = report_data.get("voting_records", [])
    voters = report_data.get("voters", [])
    stats = report_data.get("stats", {})
    integrity = report_data.get("integrity", {})
    generated_at = report_data.get("generated_at", "")
    prepared_by = (report_data.get("prepared_by") or "").strip()

    # --- CONSTANTS & COLORS ---
    PRIMARY_COLOR = colors.HexColor('#10B981')  # Emerald Green
    HEADER_BG = colors.HexColor('#D1FAE5')  # Light Green for Headers (High Visibility)
    HEADER_TEXT = colors.black  # Black Text for Headers
    ACCENT_COLOR = colors.HexColor('#F59E0B')  # Amber
    TEXT_COLOR = colors.HexColor('#1F2937')  # Dark Grey
    LIGHT_BG = colors.HexColor('#F3F4F6')  # Light Grey for boxes
    ZEBRA_BG = colors.HexColor('#F9FAFB')  # Very Light Grey for table rows

    try:
        # Create Document (Landscape A4 for better table fit)
        doc = SimpleDocTemplate(
            file_path,
            pagesize=landscape(A4),
            leftMargin=12 * mm,
            rightMargin=12 * mm,
            topMargin=22 * mm,
            bottomMargin=15 * mm,
            title=f"Report - {election.get('title', 'Election')}"
        )

        # --- STYLES ---
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            alignment=TA_LEFT,
            fontSize=24,
            textColor=TEXT_COLOR,
            leading=28,
            spaceAfter=2
        )

        cover_title_style = ParagraphStyle(
            'CoverTitle',
            parent=styles['Heading1'],
            alignment=TA_CENTER,
            fontSize=34,
            textColor=TEXT_COLOR,
            leading=40,
            spaceAfter=6,
            fontName='Helvetica-Bold'
        )

        cover_subtitle_style = ParagraphStyle(
            'CoverSubtitle',
            parent=styles['Normal'],
            alignment=TA_CENTER,
            fontSize=14,
            textColor=colors.HexColor('#374151'),
            leading=18,
            spaceAfter=4
        )

        cover_meta_style = ParagraphStyle(
            'CoverMeta',
            parent=styles['Normal'],
            alignment=TA_CENTER,
            fontSize=10,
            textColor=colors.HexColor('#6B7280'),
            leading=12,
            spaceAfter=2
        )

        meta_style = ParagraphStyle(
            'ReportMeta',
            parent=styles['Normal'],
            alignment=TA_LEFT,
            fontSize=10,
            textColor=colors.HexColor('#6B7280'),
            spaceAfter=20
        )

        h2_style = ParagraphStyle(
            'SectionH2',
            parent=styles['Heading2'],
            alignment=TA_LEFT,
            fontSize=14,
            textColor=PRIMARY_COLOR,
            spaceBefore=15,
            spaceAfter=10,
            fontName='Helvetica-Bold'
        )

        td_style = ParagraphStyle(
            'TableCell',
            parent=styles['Normal'],
            fontSize=9,
            alignment=TA_LEFT,
            textColor=TEXT_COLOR,
            leading=11,
            wordWrap='CJK'
        )

        td_center_style = ParagraphStyle(
            'TableCellCenter',
            parent=td_style,
            alignment=TA_CENTER
        )

        # Helper to create Paragraphs safely
        def p(text, style=td_style):
            if text is None: return Paragraph("-", style)
            return Paragraph(str(text), style)

        def _safe_int(v):
            try:
                return int(v or 0)
            except Exception:
                return 0

        def _safe_float(v):
            try:
                return float(v or 0.0)
            except Exception:
                return 0.0

        # Header/footer for all pages
        def _draw_header_footer(canvas, doc_obj):
            canvas.saveState()

            page_w, page_h = doc_obj.pagesize
            content_top_y = page_h - doc_obj.topMargin
            header_text_y = content_top_y + (6 * mm)
            header_line_y = content_top_y + (2.5 * mm)
            footer_text_y = doc_obj.bottomMargin - (6 * mm)
            footer_line_y = doc_obj.bottomMargin - (2.5 * mm)

            # Header
            canvas.setStrokeColor(PRIMARY_COLOR)
            canvas.setLineWidth(1)
            canvas.line(doc_obj.leftMargin, header_line_y, page_w - doc_obj.rightMargin, header_line_y)

            canvas.setFont('Helvetica', 9)
            canvas.setFillColor(colors.HexColor('#374151'))
            canvas.drawString(doc_obj.leftMargin, header_text_y, f"EduVote • {election.get('title', 'Election')}")

            canvas.setFont('Helvetica', 8)
            canvas.setFillColor(colors.HexColor('#6B7280'))
            canvas.drawRightString(page_w - doc_obj.rightMargin, header_text_y, f"Generated: {generated_at}")

            # Footer
            canvas.setStrokeColor(colors.HexColor('#E5E7EB'))
            canvas.setLineWidth(1)
            canvas.line(doc_obj.leftMargin, footer_line_y, page_w - doc_obj.rightMargin, footer_line_y)

            canvas.setFont('Helvetica', 8)
            canvas.setFillColor(colors.HexColor('#6B7280'))
            canvas.drawString(doc_obj.leftMargin, footer_text_y, "Official election report (system-generated)")
            canvas.drawRightString(page_w - doc_obj.rightMargin, footer_text_y, f"Page {doc_obj.page}")

            canvas.restoreState()

        elems = []

        # ==============================================================================
        # COVER PAGE (visually distinct)
        # ==============================================================================
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        logo_path = os.path.join(base_dir, 'Assets', 'lam.png')

        report_id = f"EV-{election.get('election_id', election.get('id', 'X'))}-{datetime.now().strftime('%Y%m%d%H%M%S')}"

        if os.path.exists(logo_path):
            try:
                elems.append(Spacer(1, 8 * mm))
                elems.append(Image(logo_path, width=45 * mm, height=45 * mm))
            except Exception:
                pass

        elems.append(Spacer(1, 6 * mm))
        elems.append(Paragraph("FULL DETAIL ELECTION REPORT", cover_title_style))
        elems.append(Paragraph(str(election.get('title', 'Election')), cover_subtitle_style))
        elems.append(Paragraph(f"Report ID: <b>{report_id}</b>", cover_meta_style))
        elems.append(Paragraph(f"Status: <b>{str(election.get('status', '')).upper()}</b>", cover_meta_style))
        elems.append(Paragraph(f"Period: {election.get('start_date', '')} to {election.get('end_date', '')}", cover_meta_style))
        if prepared_by:
            elems.append(Paragraph(f"Prepared by (Admin): <b>{prepared_by}</b>", cover_meta_style))
        # Place confidentiality notice early to ensure it stays on the cover page
        elems.append(Spacer(1, 2 * mm))
        elems.append(Paragraph(
            "Confidential: for authorized school use only.",
            ParagraphStyle('ConfTop', parent=styles['Normal'], alignment=TA_LEFT,
                           textColor=colors.HexColor('#6B7280'), fontSize=9)
        ))
        elems.append(Spacer(1, 10 * mm))

        cover_stats = [
            ["Eligible Voters", "Participants", "Turnout %", "Cast Votes", "Spoiled Votes"],
            [
                str(_safe_int(stats.get('eligible_voters'))),
                str(_safe_int(stats.get('participants'))),
                f"{_safe_float(stats.get('turnout_pct')):.1f}%",
                str(_safe_int(stats.get('cast_records'))),
                str(_safe_int(stats.get('spoiled_records'))),
            ]
        ]
        cover_table = Table(cover_stats, colWidths=[54 * mm] * 5)
        cover_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
            ('TEXTCOLOR', (0, 0), (-1, 0), HEADER_TEXT),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BACKGROUND', (0, 1), (-1, 1), LIGHT_BG),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.white),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        elems.append(cover_table)
        elems.append(Spacer(1, 10 * mm))

        if prepared_by:
            elems.append(Paragraph(f"Prepared by: <b>{prepared_by}</b>", ParagraphStyle('Sig', parent=styles['Normal'], alignment=TA_LEFT, textColor=TEXT_COLOR, fontSize=11, spaceAfter=4)))
        else:
            elems.append(Paragraph("Prepared by: ____________________________", ParagraphStyle('Sig', parent=styles['Normal'], alignment=TA_LEFT, textColor=TEXT_COLOR, fontSize=11, spaceAfter=4)))
        elems.append(Paragraph("Verified by: ____________________________", ParagraphStyle('Sig2', parent=styles['Normal'], alignment=TA_LEFT, textColor=TEXT_COLOR, fontSize=11, spaceAfter=4)))
        elems.append(Paragraph("Approved by: ____________________________", ParagraphStyle('Sig3', parent=styles['Normal'], alignment=TA_LEFT, textColor=TEXT_COLOR, fontSize=11, spaceAfter=4)))
        elems.append(PageBreak())

        # ==============================================================================
        # 1. HEADER SECTION (Summary Dashboard)
        # ==============================================================================
        header_data = [[
            Paragraph(f"Election Results: {election.get('title', 'Election')}", title_style),
            Paragraph("OFFICIAL REPORT",
                      ParagraphStyle('Badge', parent=styles['Normal'], alignment=TA_RIGHT, textColor=colors.grey,
                                     fontSize=14, fontName='Helvetica-Bold'))
        ]]
        header_table = Table(header_data, colWidths=[200 * mm, 70 * mm])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ]))
        elems.append(header_table)

        elems.append(
            Paragraph(f"Generated on: {generated_at}  |  Status: <b>{str(election.get('status', '')).upper()}</b>",
                      meta_style))

        # Green Divider Line
        elems.append(Table([['']], colWidths=[270 * mm], style=TableStyle([
            ('LINEBELOW', (0, 0), (-1, -1), 2, PRIMARY_COLOR),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ])))
        elems.append(Spacer(1, 10 * mm))

        # ==============================================================================
        # 2. SUMMARY CARDS
        # ==============================================================================
        total_votes = sum((c.get('votes') or c.get('actual_votes') or 0) for c in candidates)
        if total_votes == 0 and records:
            total_votes = sum(1 for r in records if (r.get('vote_status') or '').lower() == 'cast')
        total_candidates = len(candidates)
        participants = int(stats.get("participants") or len(voters) or 0)
        eligible_voters = int(stats.get("eligible_voters") or 0)
        spoiled_votes = int(stats.get("spoiled_records") or 0)
        turnout_pct = float(stats.get("turnout_pct") or 0.0)

        card_label_style = ParagraphStyle('CardLbl', parent=td_center_style, fontSize=9, textColor=colors.grey)
        card_val_style = ParagraphStyle('CardVal', parent=td_center_style, fontSize=20, textColor=PRIMARY_COLOR,
                                        fontName='Helvetica-Bold')

        card_data = [
            [
                p("CAST VOTES", card_label_style),
                p("SPOILED", card_label_style),
                p("PARTICIPANTS", card_label_style),
                p("TURNOUT", card_label_style),
            ],
            [
                p(str(int(stats.get("cast_records") or total_votes)), card_val_style),
                p(str(spoiled_votes), ParagraphStyle('SpoiledVal', parent=card_val_style, textColor=ACCENT_COLOR)),
                p(str(participants), card_val_style),
                p(f"{turnout_pct:.1f}%", ParagraphStyle('TurnoutVal', parent=card_val_style, textColor=ACCENT_COLOR)),
            ]
        ]

        card_table = Table(card_data, colWidths=[67 * mm] * 4)
        card_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), LIGHT_BG),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 3, colors.white),
        ]))
        elems.append(card_table)
        elems.append(Spacer(1, 10 * mm))

        # ==============================================================================
        # 3. VISUALIZATIONS (Charts)
        # ==============================================================================
        if have_mpl and (candidates or records):
            try:
                # --- Chart A: Participation overview (bar) ---
                eligible = int(stats.get("eligible_voters") or 0)
                participants = int(stats.get("participants") or len(voters) or 0)
                cast_votes = int(stats.get("cast_records") or 0)
                spoiled_votes = int(stats.get("spoiled_records") or 0)

                # Use page width for chart sizing to prevent side overlap
                page_w = float(doc.width)  # points
                col_w = page_w / 3.0

                def _img(buf, width_pts: float, aspect: float):
                    """Create an Image preserving aspect ratio (avoid stretch)."""
                    w = float(width_pts)
                    h = max(1.0, w * float(aspect))
                    return Image(buf, width=w, height=h)

                buf_overview = BytesIO()
                # Match figure aspect to the Image to avoid stretching
                fig_overview = (6.0, 4.0)
                plt.figure(figsize=fig_overview)
                cats = ["Eligible", "Participants", "Cast", "Spoiled"]
                vals = [eligible, participants, cast_votes, spoiled_votes]
                colors_bars = ['#10B981', '#34D399', '#059669', '#F59E0B']
                bars = plt.bar(cats, vals, color=colors_bars)
                plt.title('Participation Overview', fontsize=12, fontweight='bold', color='#374151')
                plt.ylabel('Count', fontsize=9, color='#374151')
                plt.grid(axis='y', linestyle='--', alpha=0.25)
                plt.gca().spines['top'].set_visible(False)
                plt.gca().spines['right'].set_visible(False)
                plt.xticks(fontsize=9)
                plt.yticks(fontsize=8)
                for b, v in zip(bars, vals):
                    plt.text(b.get_x() + b.get_width() / 2, b.get_height(), str(int(v)), ha='center', va='bottom', fontsize=8, color='#374151')
                plt.tight_layout(pad=1.0)
                plt.savefig(buf_overview, format='png', dpi=220, transparent=True, bbox_inches='tight', pad_inches=0.12)
                plt.close('all')
                buf_overview.seek(0)

                # --- Chart B: Ballot quality (donut) ---
                buf_quality = BytesIO()
                fig_quality = (5.0, 4.0)
                plt.figure(figsize=fig_quality)
                q_labels = ["Cast", "Spoiled"]
                q_vals = [max(0, cast_votes), max(0, spoiled_votes)]
                q_colors = ['#10B981', '#F59E0B']
                if sum(q_vals) > 0:
                    plt.pie(q_vals, labels=q_labels, autopct='%1.1f%%',
                            startangle=90, pctdistance=0.78,
                            colors=q_colors, wedgeprops=dict(width=0.42, edgecolor='white'))
                    plt.title('Ballot Quality', fontsize=12, fontweight='bold', color='#374151')
                else:
                    plt.text(0.5, 0.5, "No Ballot Data", ha='center', va='center')
                    plt.axis('off')
                plt.tight_layout(pad=1.0)
                plt.savefig(buf_quality, format='png', dpi=220, transparent=True, bbox_inches='tight', pad_inches=0.12)
                plt.close('all')
                buf_quality.seek(0)

                # --- Chart B: Winner votes per position (bar) ---
                from collections import defaultdict
                c_by_pos = defaultdict(list)
                for c in candidates:
                    pos_key = c.get('position_id') or c.get('position_title') or 'Unassigned'
                    c_by_pos[pos_key].append(c)

                pos_labels = []
                pos_winner_votes = []
                if positions:
                    for pos in positions:
                        pos_id = pos.get('position_id')
                        pos_title = pos.get('title') or pos.get('position_title') or 'Unassigned'
                        group = c_by_pos.get(pos_id) or []
                        if not group:
                            continue
                        top_votes = max(int(g.get('actual_votes') or 0) for g in group)
                        pos_labels.append(str(pos_title))
                        pos_winner_votes.append(top_votes)
                else:
                    # Fallback: group by text
                    by_title = defaultdict(list)
                    for c in candidates:
                        by_title[str(c.get('position_title') or c.get('position') or 'Unassigned')].append(c)
                    for title in sorted(by_title.keys()):
                        group = by_title[title]
                        top_votes = max(int(g.get('actual_votes') or 0) for g in group)
                        pos_labels.append(str(title))
                        pos_winner_votes.append(top_votes)

                buf_winners = BytesIO()
                fig_winners = (7.5, 4.0)
                plt.figure(figsize=fig_winners)
                if pos_labels and pos_winner_votes:
                    import textwrap
                    wrapped = [textwrap.fill(str(x), width=18) for x in pos_labels]
                    bars = plt.bar(wrapped, pos_winner_votes, color='#10B981')
                    plt.title('Winning Votes per Position', fontsize=12, fontweight='bold', color='#374151')
                    plt.ylabel('Votes', fontsize=9, color='#374151')
                    plt.xticks(rotation=0, ha='center', fontsize=8)
                    plt.yticks(fontsize=8)
                    plt.grid(axis='y', linestyle='--', alpha=0.25)
                    plt.gca().spines['top'].set_visible(False)
                    plt.gca().spines['right'].set_visible(False)
                    for b, v in zip(bars, pos_winner_votes):
                        plt.text(b.get_x() + b.get_width() / 2, b.get_height(), str(int(v)), ha='center', va='bottom', fontsize=7, color='#374151')
                else:
                    plt.text(0.5, 0.5, "No Position Data", ha='center', va='center')
                    plt.axis('off')
                plt.tight_layout(pad=1.0)
                plt.savefig(buf_winners, format='png', dpi=220, transparent=True, bbox_inches='tight', pad_inches=0.12)
                plt.close('all')
                buf_winners.seek(0)

                # --- Chart D: Top candidates (barh) ---
                top = sorted(
                    candidates,
                    key=lambda c: max(0, int(c.get('actual_votes') or c.get('votes') or 0)),
                    reverse=True
                )[:10]
                buf_top = BytesIO()
                fig_top = (10.5, 5.0)
                plt.figure(figsize=fig_top)
                if top:
                    labels = []
                    vals = []
                    for c in top:
                        name = str(c.get('full_name') or 'Unknown')
                        pos = str(c.get('position_title') or c.get('position') or 'Unassigned')
                        labels.append(f"{name} ({pos})")
                        vals.append(max(0, int(c.get('actual_votes') or c.get('votes') or 0)))
                    import textwrap
                    labels = [textwrap.fill(l, width=32) for l in labels]
                    bars = plt.barh(list(reversed(labels)), list(reversed(vals)), color='#10B981')
                    plt.title('Top Candidates (by Votes)', fontsize=12, fontweight='bold', color='#374151')
                    plt.xlabel('Votes', fontsize=9, color='#374151')
                    plt.xticks(fontsize=8)
                    plt.yticks(fontsize=7)
                    plt.grid(axis='x', linestyle='--', alpha=0.25)
                    plt.gca().spines['top'].set_visible(False)
                    plt.gca().spines['right'].set_visible(False)
                    for b in bars:
                        plt.text(b.get_width() + 0.2, b.get_y() + b.get_height() / 2, str(int(b.get_width())), va='center', fontsize=7, color='#374151')
                else:
                    plt.text(0.5, 0.5, "No Candidate Data", ha='center', va='center')
                    plt.axis('off')
                plt.tight_layout(pad=1.0)
                plt.savefig(buf_top, format='png', dpi=220, transparent=True, bbox_inches='tight', pad_inches=0.12)
                plt.close('all')
                buf_top.seek(0)

                # --- Chart C: Votes by rank (line) ---
                cand_votes_sorted = sorted(
                    [max(0, int(c.get('actual_votes') or c.get('votes') or 0)) for c in candidates],
                    reverse=True
                )
                buf_rank = BytesIO()
                fig_rank = (10.5, 3.6)
                plt.figure(figsize=fig_rank)
                if cand_votes_sorted:
                    x = list(range(1, len(cand_votes_sorted) + 1))
                    plt.plot(x, cand_votes_sorted, color='#10B981', marker='o', linewidth=2)
                    plt.fill_between(x, cand_votes_sorted, color='#10B981', alpha=0.12)
                    plt.title('Competitiveness (Votes by Candidate Rank)', fontsize=11, fontweight='bold', color='#374151')
                    plt.xlabel('Candidate Rank (highest to lowest)', fontsize=8)
                    plt.ylabel('Votes', fontsize=8)
                    if len(x) > 20:
                        step = max(1, len(x) // 10)
                        plt.xticks(x[::step], fontsize=7)
                    else:
                        plt.xticks(fontsize=7)
                    plt.yticks(fontsize=7)
                    plt.grid(axis='y', linestyle='--', alpha=0.25)
                    plt.gca().spines['top'].set_visible(False)
                    plt.gca().spines['right'].set_visible(False)
                else:
                    plt.text(0.5, 0.5, "No Candidate Votes", ha='center', va='center')
                    plt.axis('off')
                plt.tight_layout(pad=1.0)
                plt.savefig(buf_rank, format='png', dpi=220, transparent=True, bbox_inches='tight', pad_inches=0.12)
                plt.close('all')
                buf_rank.seek(0)

                img_overview = _img(buf_overview, col_w, fig_overview[1] / fig_overview[0])
                img_quality = _img(buf_quality, col_w, fig_quality[1] / fig_quality[0])
                img_winners = _img(buf_winners, col_w, fig_winners[1] / fig_winners[0])
                img_rank = _img(buf_rank, page_w, fig_rank[1] / fig_rank[0])
                img_top = _img(buf_top, page_w, fig_top[1] / fig_top[0])

                top_row = Table([[img_overview, img_quality, img_winners]], colWidths=[col_w, col_w, col_w])
                top_row.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))

                # Requirement: charts row on SAME page as stat cards; competitiveness on NEXT page
                elems.append(KeepTogether([
                    Paragraph("Charts & Insights", h2_style),
                    Paragraph("Turnout, ballot quality, and winners overview.", td_style),
                    Spacer(1, 2 * mm),
                    top_row,
                ]))

                elems.append(PageBreak())
                elems.append(Paragraph("Competitiveness", h2_style))
                elems.append(Paragraph("Votes by candidate rank (higher rank = more votes).", td_style))
                elems.append(Spacer(1, 3 * mm))
                elems.append(img_rank)
                elems.append(Spacer(1, 6 * mm))

                # Requirement: move the Top Candidates header to the next page
                elems.append(PageBreak())
                elems.append(Paragraph("Top Candidates", h2_style))
                elems.append(img_top)

            except Exception:
                elems.append(Paragraph("Charts unavailable for this report run.", td_style))
                elems.append(Spacer(1, 6 * mm))

        # ============================================================================== 
        # 4. EXECUTIVE SUMMARY (Eligibility + Integrity)
        # ============================================================================== 
        # Start Executive Summary on a fresh page to avoid the header being orphaned.
        elems.append(PageBreak())

        allowed_grade = election.get("allowed_grade")
        allowed_section = (election.get("allowed_section") or "").strip() or "ALL"
        eligibility_txt = f"Eligible voters: <b>{eligible_voters}</b> | Rule: Grade <b>{allowed_grade if allowed_grade is not None else 'ALL'}</b>, Section <b>{allowed_section}</b>"
        # Keep the Executive Summary header with its opening content.
        exec_intro = [
            Paragraph("2. Executive Summary", h2_style),
            Paragraph(eligibility_txt, td_style),
            Spacer(1, 4),
        ]

        integrity_rows = [
            ["CHECK", "COUNT"],
            ["Cast votes missing candidate_id", str(int(integrity.get("cast_missing_candidate") or 0))],
            ["Votes with missing user record", str(int(integrity.get("orphan_user_votes") or 0))],
            ["Votes with missing candidate record", str(int(integrity.get("orphan_candidate_votes") or 0))],
            ["Votes with missing position record", str(int(integrity.get("orphan_position_votes") or 0))],
        ]
        integ_table = Table(integrity_rows, colWidths=[200 * mm, 60 * mm], repeatRows=1)
        integ_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
            ('TEXTCOLOR', (0, 0), (-1, 0), HEADER_TEXT),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#E5E7EB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, ZEBRA_BG]),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elems.append(KeepTogether(exec_intro + [integ_table, Spacer(1, 8)]))

        # ============================================================================== 
        # 5. POSITION WINNERS + RESULTS (Grouped)
        # ============================================================================== 
        elems.append(Paragraph("3. Position Winners", h2_style))

        # Build helpers: candidates by position
        from collections import defaultdict
        c_by_pos = defaultdict(list)
        for c in candidates:
            pos_key = c.get('position_id') or c.get('position_title') or 'Unassigned'
            c_by_pos[pos_key].append(c)

        winner_rows = [["POSITION", "WINNER", "VOTES", "PCT (POSITION)"]]
        for pos in (positions or []):
            pos_id = pos.get('position_id')
            pos_title = pos.get('title') or pos.get('position_title') or 'Unassigned'
            group = c_by_pos.get(pos_id) or []
            group_sorted = sorted(group, key=lambda x: int(x.get('actual_votes') or 0), reverse=True)
            if not group_sorted:
                winner_rows.append([pos_title, "-", "0", "0.0%"])
                continue
            top_votes = int(group_sorted[0].get('actual_votes') or 0)
            tied = [g for g in group_sorted if int(g.get('actual_votes') or 0) == top_votes]
            winner_name = ", ".join([str(t.get('full_name') or '-').upper() for t in tied])
            pos_total = sum(int(g.get('actual_votes') or 0) for g in group_sorted)
            pct = (top_votes / pos_total * 100.0) if pos_total > 0 else 0.0
            if len(tied) > 1:
                winner_name = f"TIE: {winner_name}"
            winner_rows.append([pos_title, winner_name, str(top_votes), f"{pct:.1f}%"])

        # Include any candidates that are unassigned to a Position record
        if c_by_pos and positions:
            known_ids = {p.get('position_id') for p in positions}
            extras = [k for k in c_by_pos.keys() if isinstance(k, int) and k not in known_ids]
            if extras:
                for k in extras:
                    group = c_by_pos.get(k) or []
                    group_sorted = sorted(group, key=lambda x: int(x.get('actual_votes') or 0), reverse=True)
                    pos_title = str(group_sorted[0].get('position_title') or 'Unassigned')
                    top_votes = int(group_sorted[0].get('actual_votes') or 0) if group_sorted else 0
                    tied = [g for g in group_sorted if int(g.get('actual_votes') or 0) == top_votes]
                    winner_name = ", ".join([str(t.get('full_name') or '-').upper() for t in tied]) if tied else "-"
                    pos_total = sum(int(g.get('actual_votes') or 0) for g in group_sorted)
                    pct = (top_votes / pos_total * 100.0) if pos_total > 0 else 0.0
                    if len(tied) > 1:
                        winner_name = f"TIE: {winner_name}"
                    winner_rows.append([pos_title, winner_name, str(top_votes), f"{pct:.1f}%"])

        winners_table = Table(winner_rows, colWidths=[70 * mm, 120 * mm, 25 * mm, 35 * mm], repeatRows=1)
        winners_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
            ('TEXTCOLOR', (0, 0), (-1, 0), HEADER_TEXT),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#E5E7EB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, ZEBRA_BG]),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elems.append(winners_table)
        elems.append(Spacer(1, 10))

        elems.append(Paragraph("4. Detailed Results (By Position)", h2_style))

        col_widths = [18 * mm, 55 * mm, 95 * mm, 30 * mm, 35 * mm]
        headers = ["RANK", "POSITION", "CANDIDATE", "VOTES", "PCT"]

        # Render one compact table per position for readability
        def _render_position_table(position_title: str, group: list[dict]):
            group_sorted = sorted(group, key=lambda x: int(x.get('actual_votes') or 0), reverse=True)
            pos_total = sum(int(g.get('actual_votes') or 0) for g in group_sorted)
            table_data = [headers]
            for i, c in enumerate(group_sorted, 1):
                v = int(c.get('actual_votes') or 0)
                pct = (v / pos_total * 100.0) if pos_total > 0 else 0.0
                table_data.append([
                    p(f"#{i}", td_center_style),
                    p(str(position_title), td_style),
                    p(str(c.get('full_name') or 'Unknown').upper(),
                      ParagraphStyle('CandName', parent=td_style, fontName='Helvetica-Bold')),
                    p(str(v), ParagraphStyle('Votes', parent=td_center_style, fontName='Helvetica-Bold', textColor=PRIMARY_COLOR)),
                    p(f"{pct:.1f}%", td_center_style),
                ])

            t = Table(table_data, colWidths=col_widths, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
                ('TEXTCOLOR', (0, 0), (-1, 0), HEADER_TEXT),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#E5E7EB')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, ZEBRA_BG]),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elems.append(t)
            elems.append(Spacer(1, 8))

        if positions:
            for pos in positions:
                pos_id = pos.get('position_id')
                pos_title = pos.get('title') or pos.get('position_title') or 'Unassigned'
                group = c_by_pos.get(pos_id) or []
                if not group:
                    continue
                _render_position_table(pos_title, group)
        else:
            # Fallback: group by text title
            by_title = defaultdict(list)
            for c in candidates:
                by_title[str(c.get('position_title') or c.get('position') or 'Unassigned')].append(c)
            for title in sorted(by_title.keys()):
                _render_position_table(title, by_title[title])

        # ============================================================================== 
        # 6. VOTING RECORDS (Complex Table)
        # ==============================================================================
        # Avoid forcing a new page (reduces wasted whitespace). The table can split across pages.
        elems.append(Spacer(1, 8))
        elems.append(Paragraph("5. Full Voting Audit Log", h2_style))

        # Adjusted Columns for Landscape A4 (Total ~260mm usable)
        rec_col_widths = [12 * mm, 40 * mm, 22 * mm, 48 * mm, 12 * mm, 18 * mm, 35 * mm, 32 * mm, 18 * mm, 35 * mm]

        # HEADERS: Using SIMPLE STRINGS
        rec_headers = [
            "ID",
            "STUDENT NAME",
            "STUDENT ID",
            "EMAIL ADDRESS",
            "GR",
            "SEC",
            "POSITION",
            "VOTED FOR",
            "STATUS",
            "DATE VOTED"
        ]

        rec_rows = [rec_headers]

        for r in records:
            v_date = str(r.get('voted_at', ''))
            voted_for = r.get('candidate_name', '')
            if not voted_for:
                voted_for = '-' if (r.get('vote_status') or '').lower() == 'cast' else 'SPOILED'

            row = [
                p(str(r.get('record_id', '')), td_center_style),
                p(r.get('voter_name', r.get('voter_username', '')), td_style),
                p(str(r.get('voter_student_id', '')), td_style),
                p(r.get('voter_email', '') or '-', td_style),
                p(str(r.get('voter_grade', '')), td_center_style),
                p(str(r.get('voter_section', '')), td_style),
                p(r.get('position_title', '') or '-', td_style),
                p(voted_for,
                  ParagraphStyle('CandSmall', parent=td_style, textColor=colors.HexColor('#047857'))),
                p(r.get('vote_status', '').upper(), td_center_style),
                p(v_date, td_style),
            ]
            rec_rows.append(row)

        rec_table = Table(rec_rows, colWidths=rec_col_widths, repeatRows=1)
        rec_table.setStyle(TableStyle([
            # Header Row Style
            ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),  # Light Green
            ('TEXTCOLOR', (0, 0), (-1, 0), HEADER_TEXT),  # Black
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),

            # Data Row Style
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#E5E7EB')),
            ('ROWBACKGROUNDS', (1, 0), (-1, -1), [colors.white, ZEBRA_BG]),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elems.append(rec_table)

        doc.build(elems, onFirstPage=_draw_header_footer, onLaterPages=_draw_header_footer)
        return True, f"PDF report saved to: {file_path}"

    except Exception as e:
        import traceback
        traceback.print_exc()
        return False, f"Failed to generate PDF report: {e}"


def export_full_reports(election_id: int, output_path: str, prepared_by: str | None = None) -> tuple[bool, str]:
    """
    High-level export helper that creates CSV files, an Excel workbook, and a full-detail PDF.
    """
    base = os.path.splitext(output_path)[0]
    csv_entry_path = f"{base}.csv"
    excel_path = f"{base}.xlsx"
    pdf_path = f"{base}_full_detail.pdf"

    report_data = get_full_election_report_data(election_id)
    if not report_data.get("success"):
        return False, f"Failed to gather report data: {report_data.get('error')}"

    if prepared_by:
        report_data["prepared_by"] = str(prepared_by)

    # 1. Generate CSVs
    ok, msg_csv = generate_csv_report(report_data, csv_entry_path)
    if not ok:
        return False, f"CSV generation failed: {msg_csv}"

    # 2. Generate Excel
    ok_x, msg_x = generate_excel_report(report_data, excel_path)
    excel_msg = f"Excel saved: {excel_path}" if ok_x else f"Excel warning: {msg_x}"

    # 3. Generate PDF
    ok_p, msg_p = generate_pdf_report(report_data, pdf_path)
    if not ok_p: return False, f"PDF generation failed: {msg_p}"

    return True, (
        "Exports complete:\n"
        f"- PDF: {pdf_path}\n"
        f"- Excel: {excel_path} ({'ok' if ok_x else 'warning'})\n"
        f"- CSV: {csv_entry_path}\n"
        f"- Details: {msg_csv}\n"
        f"{excel_msg}"
    )


def generate_csv_report(report_data: dict, file_path: str) -> tuple[bool, str]:
    if not report_data.get("success"):
        return False, report_data.get("error", "No data available")

    election = report_data.get("election", {})
    positions = report_data.get("positions", [])
    candidates = report_data.get("candidates", [])
    records = report_data.get("voting_records", [])
    voters = report_data.get("voters", [])
    stats = report_data.get("stats", {})
    integrity = report_data.get("integrity", {})

    try:
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Section: Election Summary
            writer.writerow(["SECTION", "ELECTION SUMMARY"])
            writer.writerow(["Field", "Value"])
            writer.writerow(["Title", election.get("title", "")])
            writer.writerow(["Description", election.get("description", "")])
            writer.writerow(["Start Date", election.get("start_date", "")])
            writer.writerow(["End Date", election.get("end_date", "")])
            writer.writerow(["Status", election.get("status", "")])
            writer.writerow(["Allowed Grade", election.get("allowed_grade", "ALL")])
            writer.writerow(["Allowed Section", election.get("allowed_section", "ALL")])
            writer.writerow(["Eligible Voters", stats.get("eligible_voters", 0)])
            writer.writerow(["Participants", stats.get("participants", 0)])
            writer.writerow(["Turnout %", f"{float(stats.get('turnout_pct') or 0.0):.1f}%"])
            writer.writerow(["Cast Votes", stats.get("cast_records", 0)])
            writer.writerow(["Spoiled Votes", stats.get("spoiled_records", 0)])
            writer.writerow(["Integrity: Cast missing candidate", integrity.get("cast_missing_candidate", 0)])
            writer.writerow([])

            # Section: Positions
            writer.writerow(["SECTION", "POSITIONS"])
            writer.writerow(["position_id", "title", "display_order", "created_at"])
            for p in positions:
                writer.writerow([
                    p.get("position_id"),
                    p.get("title"),
                    p.get("display_order"),
                    p.get("created_at"),
                ])
            writer.writerow([])

            # Section: Candidates
            writer.writerow(["SECTION", "CANDIDATES"])
            writer.writerow(["rank", "candidate_id", "position", "full_name", "slogan", "email", "phone", "votes"])
            for i, c in enumerate(candidates, 1):
                writer.writerow([
                    i,
                    c.get("candidate_id"),
                    c.get("position_title") or c.get("position") or "Unassigned",
                    c.get("full_name"),
                    c.get("slogan") or "",
                    c.get("email") or "",
                    c.get("phone") or "",
                    c.get("actual_votes") or 0,
                ])
            writer.writerow([])

            # Section: Voting Records (Audit Log)
            writer.writerow(["SECTION", "VOTING RECORDS (AUDIT LOG)"])
            writer.writerow([
                "record_id", "user_id", "voter_username", "voter_name", "voter_student_id",
                "voter_email", "voter_grade", "voter_section",
                "position_id", "position_title",
                "candidate_id", "candidate_name", "status", "voted_at"
            ])
            for r in records:
                writer.writerow([
                    r.get("record_id"),
                    r.get("user_id"),
                    r.get("voter_username"),
                    r.get("voter_name"),
                    r.get("voter_student_id"),
                    r.get("voter_email"),
                    r.get("voter_grade"),
                    r.get("voter_section"),
                    r.get("position_id"),
                    r.get("position_title"),
                    r.get("candidate_id"),
                    r.get("candidate_name"),
                    r.get("vote_status"),
                    r.get("voted_at"),
                ])
            writer.writerow([])

            # Section: Participants
            writer.writerow(["SECTION", "PARTICIPANTS"])
            writer.writerow([
                "user_id", "username", "full_name", "student_id", "email",
                "grade_level", "section", "role", "user_created_at", "last_voted_at"
            ])
            for v in voters:
                writer.writerow([
                    v.get("user_id"),
                    v.get("username"),
                    v.get("full_name"),
                    v.get("student_id"),
                    v.get("email"),
                    v.get("grade_level"),
                    v.get("section"),
                    v.get("role"),
                    v.get("user_created_at"),
                    v.get("voted_at"),
                ])

        return True, f"CSV created: {os.path.basename(file_path)}"
    except Exception as e:
        return False, f"CSV Error: {e}"


def generate_excel_report(report_data: dict, file_path: str) -> tuple[bool, str]:
    if not report_data.get("success"):
        return False, report_data.get("error", "No data available")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        election = report_data.get("election", {})
        positions = report_data.get("positions", [])
        candidates = report_data.get("candidates", [])
        records = report_data.get("voting_records", [])
        voters = report_data.get("voters", [])
        stats = report_data.get("stats", {})
        integrity = report_data.get("integrity", {})

        header_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        header_font = Font(bold=True)
        wrap = Alignment(wrap_text=True, vertical="top")

        def write_table(ws, headers, rows):
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for row in rows:
                ws.append(row)
            # basic sizing
            for col_idx in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = min(45, max(12, len(str(headers[col_idx - 1])) + 2))
            for r in ws.iter_rows(min_row=2):
                for cell in r:
                    cell.alignment = wrap

        wb = Workbook()

        # Summary sheet
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Election Full Detail Report"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([])
        ws.append(["Title", election.get("title", "")])
        ws.append(["Status", election.get("status", "")])
        ws.append(["Start Date", str(election.get("start_date", ""))])
        ws.append(["End Date", str(election.get("end_date", ""))])
        ws.append(["Allowed Grade", election.get("allowed_grade", "ALL")])
        ws.append(["Allowed Section", election.get("allowed_section", "ALL")])
        ws.append(["Eligible Voters", int(stats.get("eligible_voters") or 0)])
        ws.append(["Participants", int(stats.get("participants") or 0)])
        ws.append(["Turnout %", float(stats.get("turnout_pct") or 0.0)])
        ws.append(["Cast Votes", int(stats.get("cast_records") or 0)])
        ws.append(["Spoiled Votes", int(stats.get("spoiled_records") or 0)])
        ws.append([])
        ws.append(["Integrity - Cast missing candidate", int(integrity.get("cast_missing_candidate") or 0)])
        ws.append(["Integrity - Missing user", int(integrity.get("orphan_user_votes") or 0)])
        ws.append(["Integrity - Missing candidate", int(integrity.get("orphan_candidate_votes") or 0)])
        ws.append(["Integrity - Missing position", int(integrity.get("orphan_position_votes") or 0)])
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 60

        # Positions sheet
        ws_pos = wb.create_sheet("Positions")
        write_table(
            ws_pos,
            ["position_id", "title", "display_order", "created_at"],
            [[p.get("position_id"), p.get("title"), p.get("display_order"), str(p.get("created_at"))] for p in positions],
        )

        # Candidates sheet
        ws_c = wb.create_sheet("Candidates")
        write_table(
            ws_c,
            ["candidate_id", "position", "full_name", "slogan", "email", "phone", "votes"],
            [[
                c.get("candidate_id"),
                c.get("position_title") or c.get("position") or "Unassigned",
                c.get("full_name"),
                c.get("slogan") or "",
                c.get("email") or "",
                c.get("phone") or "",
                int(c.get("actual_votes") or 0),
            ] for c in candidates],
        )

        # Voting records sheet
        ws_r = wb.create_sheet("VotingRecords")
        write_table(
            ws_r,
            [
                "record_id", "user_id", "voter_username", "voter_name", "student_id", "email",
                "grade", "section", "position", "candidate", "status", "voted_at"
            ],
            [[
                r.get("record_id"),
                r.get("user_id"),
                r.get("voter_username"),
                r.get("voter_name"),
                r.get("voter_student_id"),
                r.get("voter_email"),
                r.get("voter_grade"),
                r.get("voter_section"),
                r.get("position_title") or "",
                r.get("candidate_name") or "",
                r.get("vote_status"),
                str(r.get("voted_at")),
            ] for r in records],
        )

        # Participants sheet
        ws_v = wb.create_sheet("Participants")
        write_table(
            ws_v,
            ["user_id", "username", "full_name", "student_id", "email", "grade_level", "section", "last_voted_at"],
            [[
                v.get("user_id"),
                v.get("username"),
                v.get("full_name"),
                v.get("student_id"),
                v.get("email"),
                v.get("grade_level"),
                v.get("section"),
                str(v.get("voted_at")),
            ] for v in voters],
        )

        wb.save(file_path)
        return True, "Excel saved"
    except ImportError:
        return False, "openpyxl not installed"
    except Exception as e:
        return False, str(e)